"""T-01 单元测试:TestCase 结构体 + Excel 解析器。

覆盖:表头模糊匹配、三种步骤拆分格式、空行跳过、合并单元格、
base_url 注入、ID 列读取与缺失回退。
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from input.excel_parser import _split_items, parse_excel
from input.models import TestCase

# ── _split_items:步骤拆分(纯函数,重点测) ──────────────────────


def test_split_plain_newlines():
    text = "打开登录页\n输入用户名\n点击登录"
    assert _split_items(text) == ["打开登录页", "输入用户名", "点击登录"]


def test_split_arabic_numbered_multiline():
    text = "1. 打开登录页\n2. 输入用户名\n3. 点击登录"
    assert _split_items(text) == ["打开登录页", "输入用户名", "点击登录"]


def test_split_arabic_various_separators():
    text = "1、第一步\n2)第二步\n3）第三步\n4. 第四步"
    assert _split_items(text) == ["第一步", "第二步", "第三步", "第四步"]


def test_split_circled_multiline():
    text = "①打开\n②输入\n③提交"
    assert _split_items(text) == ["打开", "输入", "提交"]


def test_split_circled_inline_same_line():
    text = "①打开登录页②输入用户名③点击登录"
    assert _split_items(text) == ["打开登录页", "输入用户名", "点击登录"]


def test_split_arabic_inline_same_line():
    text = "1. 打开 2. 输入 3. 提交"
    assert _split_items(text) == ["打开", "输入", "提交"]


def test_split_does_not_break_decimals():
    # 小数不应被当作枚举切开
    text = "设置金额为 3.5 元后提交"
    assert _split_items(text) == ["设置金额为 3.5 元后提交"]


def test_split_drops_empty_and_whitespace_lines():
    text = "\n\n  1. 步骤一  \n\n2. 步骤二\n   \n"
    assert _split_items(text) == ["步骤一", "步骤二"]


def test_split_none_and_empty():
    assert _split_items(None) == []
    assert _split_items("") == []
    assert _split_items("   ") == []


def test_split_colon_suffix_prefix():
    text = "1：第一步\n2：第二步"
    assert _split_items(text) == ["第一步", "第二步"]


# ── parse_excel:整表解析 ───────────────────────────────────────


def _make_xlsx(tmp_path, rows, headers=None):
    """rows: list of (id, name, precond, steps, expected) tuples."""
    headers = headers or ["用例编号", "用例名称", "预置条件", "测试步骤", "预期结果"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    path = tmp_path / "cases.xlsx"
    wb.save(path)
    return path


def test_parse_basic(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [
            (
                "TC001",
                "登录成功",
                "已部署系统",
                "1. 打开登录页\n2. 输入账号\n3. 点击登录",
                "跳转到首页",
            ),
        ],
    )
    cases = parse_excel(path, base_url="http://10.0.0.1")
    assert len(cases) == 1
    c = cases[0]
    assert isinstance(c, TestCase)
    assert c.id == "TC001"
    assert c.name == "登录成功"
    assert c.preconditions == ["已部署系统"]
    assert c.steps == ["打开登录页", "输入账号", "点击登录"]
    assert c.expected == ["跳转到首页"]
    assert c.base_url == "http://10.0.0.1"  # 由参数注入,非来自 Excel


def test_parse_skips_empty_rows(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [
            ("TC001", "用例一", "", "步骤", "结果"),
            (None, None, None, None, None),  # 全空行
            ("", "", "", "", ""),  # 空字符串行
            ("TC002", "用例二", "", "步骤", "结果"),
        ],
    )
    cases = parse_excel(path)
    assert [c.id for c in cases] == ["TC001", "TC002"]


def test_parse_header_fuzzy_match(tmp_path):
    # 别名 + 多余空格
    path = _make_xlsx(
        tmp_path,
        [("X1", "用例", "前置", "做点啥", "成功")],
        headers=[" 编号 ", "用例标题", "前置条件", "操作步骤", "期望结果"],
    )
    cases = parse_excel(path)
    assert len(cases) == 1
    assert cases[0].id == "X1"
    assert cases[0].name == "用例"
    assert cases[0].preconditions == ["前置"]
    assert cases[0].steps == ["做点啥"]
    assert cases[0].expected == ["成功"]


def test_parse_id_column_used(tmp_path):
    path = _make_xlsx(tmp_path, [("BIZ-99", "用例", "", "步骤", "结果")])
    cases = parse_excel(path)
    assert cases[0].id == "BIZ-99"


def test_parse_missing_id_autogenerates(tmp_path):
    path = _make_xlsx(
        tmp_path,
        [
            ("", "用例一", "", "步骤", "结果"),
            ("", "用例二", "", "步骤", "结果"),
        ],
    )
    cases = parse_excel(path)
    assert [c.id for c in cases] == ["TC001", "TC002"]


def test_parse_missing_required_column_raises(tmp_path):
    # 缺「测试步骤」列
    path = _make_xlsx(
        tmp_path,
        [("TC001", "用例", "前置", "结果")],
        headers=["用例编号", "用例名称", "预置条件", "预期结果"],
    )
    with pytest.raises(ValueError, match="缺少必需列"):
        parse_excel(path)


def test_parse_merged_cells(tmp_path):
    # 两行用例共享一个纵向合并的「预置条件」单元格
    wb = Workbook()
    ws = wb.active
    ws.append(["用例编号", "用例名称", "预置条件", "测试步骤", "预期结果"])
    ws.append(["TC001", "用例一", "已登录", "步骤A", "结果A"])
    ws.append(["TC002", "用例二", None, "步骤B", "结果B"])
    ws.merge_cells("C2:C3")  # 预置条件纵向合并
    path = tmp_path / "merged.xlsx"
    wb.save(path)

    cases = parse_excel(path)
    assert len(cases) == 2
    assert cases[0].preconditions == ["已登录"]
    assert cases[1].preconditions == ["已登录"]  # 合并锚点值下沉


def test_parse_empty_sheet(tmp_path):
    wb = Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    assert parse_excel(path) == []
