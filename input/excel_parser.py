"""Excel 用例解析(规格 §5.1)。

布局约定(已与用户确认):
- 一行一个用例;首行为表头。
- 表头列:用例编号(ID) / 用例名称 / 预置条件 / 测试步骤 / 预期结果。
  表头做模糊匹配,容忍空格与常见别名。
- ``base_url`` 不在 Excel 内,由调用方(cli/run_case.py 或 Suite 配置)注入,
  解析器产出的 TestCase.base_url 留空。

容错(规格 §5.1):空行跳过、合并单元格按锚点取值、步骤多格式拆分。

输出:``list[TestCase]``(实现原则 3:所有来源都产出同一 TestCase 结构)。
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from .models import TestCase

logger = logging.getLogger(__name__)


# ── 表头别名(规范化后匹配:去空格、ASCII 转小写) ──────────────────
_HEADER_ALIASES: dict[str, list[str]] = {
    "id": ["用例编号", "用例id", "用例编码", "用例号", "编号", "caseid", "case_id", "id"],
    "name": ["用例名称", "用例名", "用例标题", "测试用例", "名称", "标题", "casename", "name"],
    "preconditions": ["预置条件", "前置条件", "前提条件", "预置", "precondition", "preconditions"],
    "steps": ["测试步骤", "操作步骤", "执行步骤", "步骤", "step", "steps"],
    "expected": ["预期结果", "期望结果", "预期输出", "预期", "expectedresult", "expected"],
}

# 圈号 ①..⑳
_CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"

# 行首枚举前缀:  "1." "1、" "1)" "1）" "(1)" "（1）" "①" 等(末尾允许冒号)
_PREFIX_RE = re.compile(
    rf"^\s*(?:[{_CIRCLED}]|[(（]\s*\d{{1,3}}\s*[)）]|\d{{1,3}}\s*[.、．)）:：])\s*"
)

# 行内拆分点:圈号前(总是安全),或「空白/行首 + 数字 + 分隔符 + 空白」前
# 用 (?<![\d.]) 避免把小数 "3.5" 当作枚举切开。
_INLINE_SPLIT_RE = re.compile(rf"(?=[{_CIRCLED}])" rf"|(?<![\d.])(?=\d{{1,3}}\s*[.、．)）]\s)")


def _normalize_header(value: object) -> str:
    """规范化表头:转字符串、去所有空白、ASCII 小写。"""
    return re.sub(r"\s+", "", str(value or "")).lower()


def _resolve_headers(header_cells: list[object]) -> dict[str, int]:
    """表头行 → {字段名: 列索引}。模糊匹配,首个命中列生效。"""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        norm = _normalize_header(cell)
        if not norm:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in mapping:
                continue
            if norm in aliases:
                mapping[field] = idx
                break
    return mapping


def _split_items(text: object) -> list[str]:
    """把单元格文本拆成条目列表,支持三种格式混用:

    - 纯换行
    - ``1. 2. 3.`` / ``1、`` / ``1)`` /``(1)`` 等数字枚举(行内或分行)
    - ``① ② ③`` 圈号(行内或分行)

    每条去掉枚举前缀与首尾空白,丢弃空条目。
    """
    if text is None:
        return []
    raw = str(text).replace("\r\n", "\n").replace("\r", "\n")

    items: list[str] = []
    for line in raw.split("\n"):
        # 行内可能还含多个枚举条目,再切一刀
        for piece in _INLINE_SPLIT_RE.split(line):
            cleaned = _PREFIX_RE.sub("", piece).strip()
            if cleaned:
                items.append(cleaned)
    return items


def _build_merged_resolver(ws) -> dict[tuple[int, int], object]:
    """合并单元格 → 锚点值映射 {(row, col): anchor_value}(1-based)。"""
    resolver: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                resolver[(r, c)] = anchor
    return resolver


def parse_excel(
    path: str | Path,
    *,
    base_url: str = "",
    suite_id: str | None = None,
    sheet_name: str | None = None,
) -> list[TestCase]:
    """解析 Excel,产出 ``list[TestCase]``。

    Args:
        path: Excel 文件路径(.xlsx)。
        base_url: 被测系统地址,注入每个 TestCase(Excel 内不含)。
        suite_id: 可选,归属套件 ID。
        sheet_name: 可选,指定 sheet;默认取活动 sheet。
    """
    wb = load_workbook(filename=str(path), data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    merged = _build_merged_resolver(ws)

    def cell_value(row: int, col: int) -> object:
        """col 为 0-based 字段列索引;返回合并锚点解析后的值。"""
        real_col = col + 1  # openpyxl 1-based
        if (row, real_col) in merged:
            return merged[(row, real_col)]
        return ws.cell(row=row, column=real_col).value

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = _resolve_headers(list(header_row))
    missing = [f for f in ("name", "steps") if f not in headers]
    if missing:
        wb.close()
        raise ValueError(
            f"Excel 表头缺少必需列 {missing};识别到的表头={list(header_row)}。"
            f"支持的别名见 _HEADER_ALIASES。"
        )

    cases: list[TestCase] = []
    auto_seq = 0
    # 数据行从第 2 行开始(1-based,表头是第 1 行)
    for r in range(2, ws.max_row + 1):

        def get(field: str) -> object:
            col = headers.get(field)
            return cell_value(r, col) if col is not None else None

        name_raw = get("name")
        steps_raw = get("steps")

        # 空行跳过:名称与步骤都为空视为空行
        if not str(name_raw or "").strip() and not str(steps_raw or "").strip():
            continue

        auto_seq += 1
        case_id = str(get("id") or "").strip()
        if not case_id:
            case_id = f"TC{auto_seq:03d}"
            logger.warning("第 %d 行缺少用例 ID,自动生成 %s", r, case_id)

        cases.append(
            TestCase(
                id=case_id,
                name=str(name_raw or "").strip(),
                preconditions=_split_items(get("preconditions")),
                steps=_split_items(steps_raw),
                expected=_split_items(get("expected")),
                base_url=base_url,
                suite_id=suite_id,
            )
        )

    wb.close()
    return cases
